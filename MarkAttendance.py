import face_recognition
import cv2
import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import *
import os
from datetime import *

def MarkAttendance():
    # ENCODING THE KONWN FACES
    cur=os.getcwd()
    newdir=cur+"\\"+"photos"
    list_file = newdir
    known_face_encodings = []
    known_face_names = []
    for img in os.listdir(list_file):
        person_image = face_recognition.load_image_file(f'{list_file}\{img}')
        person_encodings = face_recognition.face_encodings(person_image)[0]
        known_face_encodings.append(person_encodings)
        known_face_names.append(os.path.splitext(img)[0])
        students = known_face_names.copy()

    # CREATING THE EXCEL FILE
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    file=str(current_date)+'.xlsx'
    direc=os.getcwd()
    if file not in os.listdir(direc):
        workbook = xlsxwriter.Workbook(f'{current_date}.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'NAME')
        worksheet.write('B1', 'INTIME')
        worksheet.write('C1', 'OUTTIME')
        workbook.close()
    else:
        pass    

    video_capture = cv2.VideoCapture(0)
    while True:
        _,Frame = video_capture.read()
        Frame = cv2.resize(Frame,(0,0),fx=1.5,fy=1.5)

        # FINDING THE BEST MATCH INDEX
        unknown_face_location = face_recognition.face_locations(Frame)
        unknown_face_encoding = face_recognition.face_encodings(Frame,unknown_face_location)
        for faceloc ,face_encode in zip(unknown_face_location,unknown_face_encoding):
            matches = face_recognition.compare_faces(known_face_encodings,face_encode)
            face_distance = face_recognition.face_distance(known_face_encodings,face_encode)
            face_match_percentage=(1-face_distance)*100
            face_match_percentage=np.round(face_match_percentage,2)
            best_match_index = np.argmax(face_match_percentage)
            name = " "    
            if matches[best_match_index]:
                name = known_face_names[best_match_index]

            # CREATING RECTANGLE
            x1,y1,x2,y2 = faceloc[3],faceloc[0],faceloc[1],faceloc[2]
            cv2.rectangle(Frame,(x1,y1),(x2,y2),(0,255,0),thickness=2)
            cv2.rectangle(Frame,(x1,y2),(x2,y2+35),(0,255,0),cv2.FILLED)
            cv2.putText(Frame,name,(x1+10,y2+25),cv2.FONT_HERSHEY_COMPLEX,1,(0,0,0),thickness=2)

            # MARKING ATTENDENCE
            if name in known_face_names:
                if name in students:
                    students.remove(name)
                    current_time = now.strftime("%H-%M-%S")
                    df = pd.ExcelFile(f'{current_date}.xlsx').parse("Sheet1")
                    x=[]
                    for i in range(len(df)):
                        x.append(df["NAME"][i])    
                    if name in x:    
                        wb = load_workbook(f'{current_date}.xlsx')
                        ws = wb['Sheet1']
                        index=x.index(name)
                        cell="C"+str(index+2)
                        ws[cell]=current_time    
                        wb.save(f'{current_date}.xlsx')
                    else:
                        wb = load_workbook(f'{current_date}.xlsx')
                        ws = wb['Sheet1']
                        ws.append([name,current_time,None])
                        wb.save(f'{current_date}.xlsx')

        
        cv2.imshow("Attendence System", Frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    video_capture.release()
    cv2.destroyAllWindows()